import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# aggiungi che se l'ordine è cancellato, la cella si colora di rosso
path = "Order_Status_2024OK.xlsx"

def formattingStatusOrders():
    wb = openpyxl.load_workbook(path)

    ws = wb.active

    cellafill = PatternFill(start_color="0080ccbb", end_color="0080ccbb", fill_type="solid")
    for riga in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if riga[0].row % 2 == 0:
            for cella in riga:
                cella.fill = cellafill 

    lookerPetrol = PatternFill(start_color="000d5d63", end_color="000d5d5d", fill_type="solid")
    font_white = Font(color="00ffffff")

    # coloro l'intestazione
    for colonna in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cella in colonna:
            cella.fill = lookerPetrol
            cella.font = font_white
    
    # creo i colori per riempire le celle della colonna status in base allo stato dell'ordine (spedito, pending, cancellato)
    cancel_order = PatternFill(start_color="00ff7f50", end_color="00ff7f50", fill_type="solid")
    shipped_order = PatternFill(start_color="007fff00", end_color="007fff00", fill_type="solid")
    pending_order = PatternFill(start_color="00c0c0c0", end_color="00c0c0c0", fill_type="solid")
           
    # inserisco i tre colori nelle rispettive celle       
    for riga in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
        for cella in riga:
            if cella.value == "Cancel":
                cella.fill = cancel_order
            elif not cella.value:
                cella.fill = pending_order
            else:
                cella.fill = shipped_order

    # centro tutti i valori delle celle
    for riga in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cella in riga:
            cella.alignment = Alignment(horizontal="center")
            
    ws.auto_filter.ref = ws.dimensions
    
    wb.save("G:\\Il mio Drive\\Order_Status_2024.xlsx")
    wb.save("Order_Status_2024.xlsx") # FILE LOCALE

if __name__ == "__main__":
    formattingStatusOrders()
    print("File Formattato")